# This script was created to automate the task of transforming PeakSimple LOG into comprehensible spreadsheets
# Instruments: GC-FID
# Software: works with Peak Simple and MS Excel
# Authors: Davi de Ferreyro Monticelli, iREACH group (University of British Columbia)
# Date: 2024-04-22
# Version: 1.0.0

# How it works: The script will read the Peak Simple LOG file and search for groups of files that belong
#               to a single run. It will then organize the results for search groups by creating
#               individual spreadsheets and naming them comprehensively.

#######################################################################################################################
# NOTE: This script works for LOG files in which the naming convention of original .CHR files is:
#       "Terpenes_Sampling_Zmin_XX_XX_XX_CCC-CCC_STARTXXXX_SXX", where X are numbers,
#       XX_XX_XX is the sampling date, STARTXXXX the start time, and SXX, the sample #, and
#       C are characters. If the naming convention differs, please modify the script
#       accordingly. Z is the trapping time in minutes.

# NOTE: This script was originally designed for organizing the data of 22 terpene species (from calibration)
#       Therefore it modifies the log file to include a heading (column names) such as:
#       "'filename' 'date 'hour' 'component t1' 'retention t1' 'area t1' 'internal t1' 'norm_area t1' (...)
#       'component t22' 'retention t22' 'area t22' 'internal t22' 'norm_area 22'"
#       If you are working with different species, please modify where appropriate.

# NOTE: Use the file path and name of your machine/preference.

# NOTE: Adjust trapping times and temperature cycle time accordingly to your runs (4th column).

# NOTE: This script uses the following conversion from ug/ml to:
#       (ppb) = (I*C)*(10^(-9))*(1/M)*(1/(F*T))*(1000)*(1/(4.09*(10^(-8))))*(1000)
#       where I is the 'internal' value in (ng/uL)
#             C is the injected volume when calibrating (in this case 1uL)
#             M is the molar mass of the compounds (please modify dictionary)
#             F is the volume flow (intake) in (mL/min)
#             T is the trapping time in minutes
#######################################################################################################################

import pandas as pd
import numpy as np
import re
import os
from openpyxl import Workbook

import warnings

# Get the current user's username
username = os.getlogin()

# Filter out FutureWarnings
warnings.filterwarnings("ignore", category=FutureWarning)

csv_file_path = f'C:\\Users\\{username}\\PycharmProjects\\GC-FID\\PeakSimple LOG\\'
file_name = 'PeakSimple_Log.csv'
df = pd.read_csv(csv_file_path+file_name, delimiter=',')

print("PEAK SIMPLE LOG FILE:")
df.columns = ['filename', 'date', 'hour', 'component t1', 'retention t1', 'area t1', 'internal t1', 'norm_area t1',
              'component t2', 'retention t2', 'area t2', 'internal t2', 'norm_area t2',
              'component t3', 'retention t3', 'area t3', 'internal t3', 'norm_area t3',
              'component t4', 'retention t4', 'area t4', 'internal t4', 'norm_area t4',
              'component t5', 'retention t5', 'area t5', 'internal t5', 'norm_area t5',
              'component t6', 'retention t6', 'area t6', 'internal t6', 'norm_area t6',
              'component t7', 'retention t7', 'area t7', 'internal t7', 'norm_area t7',
              'component t8', 'retention t8', 'area t8', 'internal t8', 'norm_area t8',
              'component t9', 'retention t9', 'area t9', 'internal t9', 'norm_area t9',
              'component t10', 'retention t10', 'area t10', 'internal t10', 'norm_area t10',
              'component t11', 'retention t11', 'area t11', 'internal t11', 'norm_area t11',
              'component t12', 'retention t12', 'area t12', 'internal t12', 'norm_area t12',
              'component t13', 'retention t13', 'area t13', 'internal t13', 'norm_area t13',
              'component t14', 'retention t14', 'area t14', 'internal t14', 'norm_area t14',
              'component t15', 'retention t15', 'area t15', 'internal t15', 'norm_area t15',
              'component t16', 'retention t16', 'area t16', 'internal t16', 'norm_area t16',
              'component t17', 'retention t17', 'area t17', 'internal t17', 'norm_area t17',
              'component t18', 'retention t18', 'area t18', 'internal t18', 'norm_area t18',
              'component t19', 'retention t19', 'area t19', 'internal t19', 'norm_area t19',
              'component t20', 'retention t20', 'area t20', 'internal t20', 'norm_area t20',
              'component t21', 'retention t21', 'area t21', 'internal t21', 'norm_area t21',
              'component t22', 'retention t22', 'area t22', 'internal t22', 'norm_area t22']
print(df)
print("")

# Dictionary of terpenes molar masses:
terp_M = {
          'a-pinene': 136.24,
          'camphene': 136.24,
          'b-pinene': 136.24,
          'b-myrcene': 136.24,
          'delta-3-carene': 136.24,
          'a-terpinene': 136.24,
          'd-limonene': 136.24,
          'p-cymene': 134.21,
          'eucalyptol': 154.25,
          'ocimene-2': 136.24,
          'gamma-terpinene': 136.24,
          'terpinolene': 136.24,
          'linalool': 154.25,
          'isopulegol': 154.25,
          'geraniol': 154.25,
          'b-caryophyllene': 204.36,
          'a-humulene': 204.36,
          'cis-nerolidol': 222.37,
          'trans-nerolidol': 222.37,
          'caryophyllene-oxide':  220.35,
          'guaiol': 222.37,
          'a-bisabol': 284.70
         }

# Function to extract relevant information from the filename
def extract_info(filename):
    # Define the regular expression pattern
    pattern = r'^Terpenes_Sampling_\d+min_\d{2}_\d{2}_\d{2}_[A-Za-z]+-[A-Za-z]+_START\d+_'
    # Use re.match to find the match at the beginning of the string
    match = re.match(pattern, filename)
    core_string = match.group()

    # Extract room name and start time
    match = re.search(r'Terpenes_Sampling_(\d+)min_(\d+_\d+_\d+)_(\w+-\w+)_START\d+_S(\d+)\.CHR', filename)
    trapping_t = match.group(1)
    room_date = match.group(2)
    room = match.group(3)
    sample_number = match.group(4)

    return core_string, trapping_t, room_date, room, sample_number

i = 0
filename = df.iloc[i, 0]
wb = Workbook()
while i < len(df):
    # Get general info of room group
    filename = df.iloc[i, 0]
    the_core_string, the_trapping_t, the_date, the_room, the_sample_number = extract_info(filename)

    while the_core_string == the_core_string:
        if i < len(df):
            # Create empty dataframe with 6 columns and 25 rows to print data on
            empty_df = pd.DataFrame(np.nan, index=range(25), columns=range(6))
            filename = df.iloc[i, 0]
            the_core_string, the_trapping_t, the_date, the_room, the_sample_number = extract_info(filename)
            the_sample_time = pd.to_datetime(df.iloc[i, 1] + ' ' + df.iloc[i, 2])
            print("Core string is: ", the_core_string)
            print("Trapping time is: ", the_trapping_t)
            print("Room is: ", the_room)
            print("Sample # is: ", the_sample_number)
            print("Sample datetime is:", the_sample_time)
            print("...")
            print("Adding info to spreadsheet")
            print("...")

            # Filling dataframe by columns

            # 1st Column:
            empty_df.at[0, 0] = "Room"
            empty_df.at[1, 0] = "Time"
            empty_df.at[2, 0] = "Component"
            the_terpenes = [value for value in df.iloc[i, 3:].tolist() if isinstance(value, str)]
            for t in range(0, len(the_terpenes)):
                empty_df.at[t+3, 0] = the_terpenes[t]

            # 2nd Column:
            empty_df.at[0, 1] = the_room
            empty_df.at[1, 1] = df.iloc[i, 2]
            empty_df.at[2, 1] = "Retention"
            retention_list = []
            n_col = df.shape[1]
            for r in range(3, n_col, 5):
                # Extract the first value after the strings
                retention_list.append(df.iloc[i, r + 1])
            for r in range(0, len(retention_list)):
                empty_df.at[3+r, 1] = retention_list[r]

            # 3rd Column:
            empty_df.at[0, 2] = np.nan
            empty_df.at[1, 2] = np.nan
            empty_df.at[2, 2] = "Area"
            area_list = []
            n_col = df.shape[1]
            for r in range(3, n_col, 5):
                # Extract the first value after the strings
                area_list.append(df.iloc[i, r + 2])
            for r in range(0, len(area_list)):
                empty_df.at[3+r, 2] = area_list[r]

            # 4th Column:
            if int(the_trapping_t) == 1:
                empty_df.at[0, 3] = "1:32:00"
            else:
                empty_df.at[0, 3] = "1:42:00"
            empty_df.at[1, 3] = the_trapping_t
            empty_df.at[2, 3] = "Internal"
            internal_list = []
            n_col = df.shape[1]
            for r in range(3, n_col, 5):
                # Extract the first value after the strings
                internal_list.append(df.iloc[i, r + 3])
            for r in range(0, len(internal_list)):
                empty_df.at[3+r, 3] = internal_list[r]

            # 5th Column:
            empty_df.at[0, 4] = "Temperature time"
            empty_df.at[1, 4] = "Trapping minutes"
            empty_df.at[2, 4] = "Norm area %"
            normarea_list = []
            n_col = df.shape[1]
            for r in range(3, n_col, 5):
                # Extract the first value after the strings
                normarea_list.append(df.iloc[i, r + 4])
            for r in range(0, len(normarea_list)):
                empty_df.at[3+r, 4] = normarea_list[r]

            # 6th Column:
            empty_df.at[0, 5] = np.nan
            empty_df.at[1, 5] = np.nan
            empty_df.at[2, 5] = "Units"
            trap_t = float(the_trapping_t)
            for r in range(0, len(normarea_list)):
                terpene = empty_df.at[3+r, 0]
                molar_mass = terp_M[terpene]
                empty_df.at[3+r, 5] = (empty_df.at[r+3, 3]*1)*(10**(-9))*(1/molar_mass)*(1/(100*trap_t))*1000*(1/(4.09*(10**(-8))))*1000

            print("")
            print("Sample Dataframe:")
            print(empty_df)
            print("")

            tab_name = f"S{the_sample_number}"
            ws = wb.create_sheet(title=tab_name)

            for r_idx, row in enumerate(empty_df.iterrows(), start=1):
                for c_idx, value in enumerate(row[1], start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Save the workbook
            # Remove the default "Sheet" tab
            if "Sheet" in wb.sheetnames:
                default_sheet = wb["Sheet"]
                wb.remove(default_sheet)
            else:
                wb.save(csv_file_path+f"{the_room}_{the_date}.xlsx")

            i += 1  # Update i
            if i < len(df):
                test_filename = df.iloc[i, 0]
                test_core_string, test_trapping_t, test_date, test_room, test_sample_number = extract_info(test_filename)
                if test_core_string != the_core_string:
                    wb = Workbook()
            else:
                exit()

        else:
            exit()
